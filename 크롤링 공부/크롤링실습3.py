from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl #엑셀 모듈 추가

customService = Service(ChromeDriverManager().install())
customOption = Options()
browser = webdriver.Chrome(service = customService, options = customOption)

URL = 'https://vibe.naver.com/chart/total'
browser.get(URL)
browser.implicitly_wait(10)

#엑셀 모듈에서 파일 -> 시트 객체 생성
xlsxFile = openpyxl.Workbook()
xlsxSheet = xlsxFile.active
xlsxSheet.cell(row=1,column=1).value='rank'
xlsxSheet.cell(row=1,column=2).value='title'
xlsxSheet.cell(row=1,column=3).value='artist'

#팝업창 X 클릭 코드
browser.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div/a[2]').click()

#TOP100 반복
for i in range(1, 101, 1):
    rank = browser.find_element(By.XPATH, f'//*[@id="content"]/div[4]/div[2]/div/table/tbody/tr[{i}]/td[3]/span').text
    title = browser.find_element(By.XPATH, f'//*[@id="content"]/div[4]/div[2]/div/table/tbody/tr[{i}]/td[4]/div[1]/span/a').text
    artist = browser.find_element(By.XPATH, f'//*[@id="content"]/div[4]/div[2]/div/table/tbody/tr[{i}]/td[4]/div[2]/span[1]/span/a/span').text

    print(rank, title, artist)

    #셀에 넣는 코드
    xlsxSheet.cell(row = i+1, column = 1).value = rank
    xlsxSheet.cell(row = i+1, column = 2).value = title
    xlsxSheet.cell(row = i+1, column = 3).value = artist
    
#저장
xlsxFile.save('C:/박규영/2023크롤링공부/크롤링(개발자유머)/음악차트top100.xlsx')